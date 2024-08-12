from openpyxl import Workbook, load_workbook


def config():
    try:
        wb = load_workbook(filename='campo.xlsx')
        config = wb['config']
    except:
        wb = Workbook()
        config = wb.create_sheet('config')
        config.cell(column=1, row=1, value="linha")
        config.cell(column=2, row=1, value=5)
        config.cell(column=1, row=2, value="coluna")
        config.cell(column=2, row=2, value=5)

    linhas = config.cell(column=2, row=1).value
    colunas = config.cell(column=2, row=2).value

    wb.save('campo.xlsx')

    return linhas, colunas



def criarTabuleiro(quantLinha, quantColuna):
    for i in range(int(quantLinha)):
        for j in range(0, int(quantColuna)):
            print('{:5}'.format('[ ]') , end='')
        print()

# quantColuna = input('Quantas colunas: ')
# quantLinha = input('Quantas Linhas: ')

# criarTabuleiro(quantLinha, quantColuna)


linha, coluna = config()